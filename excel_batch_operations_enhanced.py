
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, time
import argparse
import re

EXCEL_FILE = "SchoolNurse_HealthLog.xlsx"
STUDENT_DATA_SHEET = "Student Data"

# Define column headers based on excel_design_plan.md
# This list must match the order of columns in your Excel sheet
COLUMN_HEADERS = [
    "Patient ID", "Full Name", "Date of Birth", "Age", "Gender", "Grade/Year Level",
    "House/Section", "Parent/Guardian Primary Contact Name", "Parent/Guardian Primary Contact Number",
    "Parent/Guardian Secondary Contact Name", "Parent/Guardian Secondary Contact Number",
    "Emergency Contact Name", "Emergency Contact Number", "Student's Class/Homeroom Teacher",
    "Academic Year", "Academic Term", "Date of Visit", "Time of Visit", "Brought in by",
    "Nurse Name/ID", "Visit Reason Category", "Severity Level", "Temperature (°C)",
    "Pulse (bpm)", "Respiratory rate (cpm)", "Oxygen saturation (%)", "Blood pressure (mmHg)",
    "Presenting Complaint(s)", "Other Complaint Details", "Background to Presenting Complaint(s)",
    "Past Medical History", "Known Allergies", "Current Medications", "Special Medical Needs Flag",
    "Chronic Conditions Alert", "Nurse Observations", "Interventions Provided",
    "Medications Administered (during visit)", "Next Step(s)", "Other Next Step Details",
    "Referral Type", "Follow-up Date", "Date of Admission (Sick Bay)", "Time of Admission (Sick Bay)",
    "Student's Condition on Admission (Sick Bay)", "Plan of Care (Sick Bay)", "Time of Discharge",
    "Student's Condition at Discharge", "Discharge Instructions", "Return to Class Time",
    "Parent Notification (Yes/No)", "Parent Notification Time", "Incident Report Required (Yes/No)",
    "Notes/Comments"
]

# Define fields that require specific data types or validation
DATE_FIELDS = [
    "Date of Birth", "Date of Visit", "Follow-up Date", "Date of Admission (Sick Bay)"
]
TIME_FIELDS = [
    "Time of Visit", "Time of Admission (Sick Bay)", "Time of Discharge", "Return to Class Time", "Parent Notification Time"
]
NUMERIC_FIELDS = [
    "Age", "Parent/Guardian Primary Contact Number", "Parent/Guardian Secondary Contact Number",
    "Emergency Contact Number", "Temperature (°C)", "Pulse (bpm)", "Respiratory rate (cpm)",
    "Oxygen saturation (%)"
]
BOOLEAN_FIELDS = [
    "Special Medical Needs Flag", "Chronic Conditions Alert", "Parent Notification (Yes/No)",
    "Incident Report Required (Yes/No)"
]
REQUIRED_FIELDS = [
    "Full Name", "Date of Visit", "Time of Visit", "Nurse Name/ID"
]

def generate_patient_id():
    """Generates a simple patient ID based on current timestamp."""
    return "SID" + datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]

def _convert_to_excel_value(field_name, value):
    """Converts value to appropriate type for Excel, handling dates/times/booleans."""
    if value is None or value == "":
        return None

    if field_name in DATE_FIELDS:
        try:
            return datetime.strptime(str(value).split(' ')[0], "%Y-%m-%d").date() # Only date part
        except ValueError:
            return str(value) # Keep as string if conversion fails
    elif field_name in TIME_FIELDS:
        try:
            # Handle various time formats, including 'HH:MM AM/PM'
            if 'AM' in str(value).upper() or 'PM' in str(value).upper():
                return datetime.strptime(str(value), "%I:%M %p").time()
            else:
                return datetime.strptime(str(value), "%H:%M").time()
        except ValueError:
            return str(value) # Keep as string if conversion fails
    elif field_name in NUMERIC_FIELDS:
        try:
            return float(value) if '.' in str(value) else int(value)
        except ValueError:
            return str(value) # Keep as string if conversion fails
    elif field_name in BOOLEAN_FIELDS:
        return str(value).lower() == 'yes' or str(value).lower() == 'true'
    else:
        return value

def initialize_excel_file():
    """Initializes the Excel file with headers if it doesn't exist."""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active) # Remove default sheet

    if STUDENT_DATA_SHEET not in workbook.sheetnames:
        sheet = workbook.create_sheet(STUDENT_DATA_SHEET)
        sheet.append(COLUMN_HEADERS)
        # Auto-adjust column widths
        for i, header in enumerate(COLUMN_HEADERS):
            sheet.column_dimensions[get_column_letter(i + 1)].width = max(len(header), 20)
    else:
        sheet = workbook[STUDENT_DATA_SHEET]
        # Basic check if headers match
        current_headers = [cell.value for cell in sheet[1]]
        if current_headers != COLUMN_HEADERS:
            print("Warning: Headers in existing Excel file do not exactly match expected. Please verify your Excel file structure.")

    try:
        workbook.save(EXCEL_FILE)
        print(f"Excel file \'{EXCEL_FILE}\' initialized/verified.")
    except Exception as e:
        print(f"Error saving Excel file during initialization: {e}")

def validate_record(record_data):
    """Performs basic validation on record data."""
    errors = []
    for field in REQUIRED_FIELDS:
        if field not in record_data or not record_data[field]:
            errors.append(f"'{field}' is a required field.")

    for field in DATE_FIELDS:
        if field in record_data and record_data[field]:
            try:
                datetime.strptime(str(record_data[field]).split(' ')[0], "%Y-%m-%d")
            except ValueError:
                errors.append(f"'{field}' must be in YYYY-MM-DD format.")

    for field in TIME_FIELDS:
        if field in record_data and record_data[field]:
            try:
                if 'AM' in str(record_data[field]).upper() or 'PM' in str(record_data[field]).upper():
                    datetime.strptime(str(record_data[field]), "%I:%M %p")
                else:
                    datetime.strptime(str(record_data[field]), "%H:%M")
            except ValueError:
                errors.append(f"'{field}' must be in HH:MM or HH:MM AM/PM format.")

    for field in NUMERIC_FIELDS:
        if field in record_data and record_data[field]:
            try:
                float(record_data[field])
            except ValueError:
                errors.append(f"'{field}' must be a numeric value.")

    return errors

def add_record(record_data):
    """Adds a new record to the Excel sheet.
    record_data should be a dictionary where keys are column headers.
    """
    errors = validate_record(record_data)
    if errors:
        print("Validation errors:")
        for error in errors:
            print(f"- {error}")
        return False

    initialize_excel_file()
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook[STUDENT_DATA_SHEET]

        # Generate Patient ID if not provided
        if "Patient ID" not in record_data or not record_data["Patient ID"]:
            record_data["Patient ID"] = generate_patient_id()

        # Create a list of values in the correct order of columns, converting types
        row_values = [_convert_to_excel_value(header, record_data.get(header, "")) for header in COLUMN_HEADERS]
        sheet.append(row_values)
        workbook.save(EXCEL_FILE)
        print(f"Record for \'{record_data.get('Full Name', 'N/A')}\' added successfully.")
        return True
    except Exception as e:
        print(f"Error adding record: {e}")
        return False

def get_all_records():
    """Reads all records from the Excel sheet and returns them as a list of dictionaries."""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook[STUDENT_DATA_SHEET]
    except FileNotFoundError:
        print(f"Error: Excel file \'{EXCEL_FILE}\' not found. Please initialize it first.")
        return []
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

    records = []
    headers = [cell.value for cell in sheet[1]] # Get headers from the first row

    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, header in enumerate(headers):
            record[header] = row[i]
        records.append(record)
    return records

def find_records(search_criteria, exact_match=False):
    """Finds records matching the search criteria.
    search_criteria is a dictionary, e.g., {'Full Name': 'John Doe', 'Grade/Year Level': 'Year 5'}
    exact_match: if True, values must match exactly; otherwise, partial and case-insensitive match.
    """
    all_records = get_all_records()
    found = []
    for record in all_records:
        match = True
        for key, value in search_criteria.items():
            if key not in record or record[key] is None:
                match = False
                break

            record_value = str(record[key])
            search_value = str(value)

            if exact_match:
                if record_value.lower() != search_value.lower():
                    match = False
                    break
            else:
                if search_value.lower() not in record_value.lower():
                    match = False
                    break
        if match:
            found.append(record)
    return found

def update_record(patient_id, updates):
    """Updates an existing record based on Patient ID.
    updates is a dictionary of {column_header: new_value}.
    """
    initialize_excel_file()
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook[STUDENT_DATA_SHEET]

        headers = [cell.value for cell in sheet[1]]
        if "Patient ID" not in headers:
            print("Error: 'Patient ID' column not found in Excel file.")
            return False
        patient_id_col_index = headers.index("Patient ID")

        updated_count = 0
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_idx, column=patient_id_col_index + 1).value) == str(patient_id):
                for header, new_value in updates.items():
                    if header in headers:
                        col_idx = headers.index(header)
                        sheet.cell(row=row_idx, column=col_idx + 1).value = _convert_to_excel_value(header, new_value)
                updated_count += 1
                break # Assuming Patient ID is unique

        workbook.save(EXCEL_FILE)
        if updated_count > 0:
            print(f"Record for Patient ID \'{patient_id}\' updated successfully.")
            return True
        else:
            print(f"No record found for Patient ID \'{patient_id}\' to update.")
            return False
    except Exception as e:
        print(f"Error updating record: {e}")
        return False

def delete_record(patient_id):
    """Deletes a record based on Patient ID."""
    initialize_excel_file()
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook[STUDENT_DATA_SHEET]

        headers = [cell.value for cell in sheet[1]]
        if "Patient ID" not in headers:
            print("Error: 'Patient ID' column not found in Excel file.")
            return False
        patient_id_col_index = headers.index("Patient ID")

        deleted_count = 0
        for row_idx in range(sheet.max_row, 1, -1): # Iterate backwards to avoid issues with row deletion
            if str(sheet.cell(row=row_idx, column=patient_id_col_index + 1).value) == str(patient_id):
                sheet.delete_rows(row_idx)
                deleted_count += 1
                break # Assuming Patient ID is unique

        workbook.save(EXCEL_FILE)
        if deleted_count > 0:
            print(f"Record for Patient ID \'{patient_id}\' deleted successfully.")
            return True
        else:
            print(f"No record found for Patient ID \'{patient_id}\' to delete.")
            return False
    except Exception as e:
        print(f"Error deleting record: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description="School Nurse Health Log - Excel Batch Operations.")
    parser.add_argument("command", choices=["add", "list", "find", "update", "delete", "init"], help="Command to execute")
    parser.add_argument("--data", type=str, help="JSON string of data for 'add' or 'update' command. E.g., '{\"Full Name\": \"John Doe\", \"Age\": 10}'")
    parser.add_argument("--patient_id", type=str, help="Patient ID for 'update' or 'delete' command")
    parser.add_argument("--exact_match", action="store_true", help="For 'find' command, perform exact match (case-insensitive)")

    args = parser.parse_args()

    if args.command == "init":
        initialize_excel_file()
    elif args.command == "add":
        if not args.data:
            print("Error: --data argument is required for 'add' command.")
            return
        try:
            data = eval(args.data) # Using eval for simplicity, but json.loads is safer for real apps
            add_record(data)
        except Exception as e:
            print(f"Invalid data format: {e}")
    elif args.command == "list":
        records = get_all_records()
        if records:
            for record in records:
                print(record)
        else:
            print("No records found.")
    elif args.command == "find":
        if not args.data:
            print("Error: --data argument is required for 'find' command.")
            return
        try:
            criteria = eval(args.data)
            found_records = find_records(criteria, args.exact_match)
            if found_records:
                for record in found_records:
                    print(record)
            else:
                print("No records found matching criteria.")
        except Exception as e:
            print(f"Invalid criteria format: {e}")
    elif args.command == "update":
        if not args.patient_id or not args.data:
            print("Error: --patient_id and --data arguments are required for 'update' command.")
            return
        try:
            updates = eval(args.data)
            update_record(args.patient_id, updates)
        except Exception as e:
            print(f"Invalid data format: {e}")
    elif args.command == "delete":
        if not args.patient_id:
            print("Error: --patient_id argument is required for 'delete' command.")
            return
        delete_record(args.patient_id)

if __name__ == "__main__":
    main()



