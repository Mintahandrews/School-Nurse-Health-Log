import pandas as pd
import openpyxl
from datetime import datetime, date, time
import os
import re
from flask import current_app
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_excel_headers():
    """Return the list of column headers for the Excel file"""
    return [
        # Identification
        'Patient ID',
        
        # Student Demographics
        'Full Name', 'Date of Birth', 'Age', 'Gender', 'Grade/Year Level',
        
        # Contact Information
        'Parent/Guardian Name', 'Parent/Guardian Phone', 
        'Emergency Contact Name', 'Emergency Contact Phone',
        
        # Visit Information
        'Academic Year', 'Academic Term', 'Date of Visit', 'Time of Visit', 
        'Brought In By', 'Nurse Name', 'Visit Reason Category', 'Severity Level', 'Visit Details',
        
        # Vital Signs
        'Temperature (°C)', 'Pulse (bpm)', 'Respiratory Rate (cpm)', 
        'Oxygen Saturation (%)', 'Blood Pressure (mmHg)',
        
        # Presenting Complaints
        'Presenting Complaint(s)', 'Other Complaint Details', 'Complaint Background',
        
        # Medical History
        'Past Medical History', 'Known Allergies', 'Current Medications',
        'Special Medical Needs', 'Chronic Conditions',
        
        # Assessment & Care
        'Nurse Observations', 'Interventions Provided', 'Medications Administered',
        'Next Step(s)', 'Other Next Step Details', 'Referral Type', 'Follow-up Date',
        
        # Admission to Sick Bay
        'Admission Date', 'Admission Time', 'Condition on Admission', 'Plan of Care',
        
        # Discharge Information
        'Discharge Time', 'Condition at Discharge', 'Discharge Instructions',
        'Return to Class Time', 'Parent Notified', 'Parent Notification Time',
        'Incident Report Required',
        
        # System
        'Notes', 'Created At', 'Updated At'
    ]

def initialize_excel_file(excel_file=None):
    """Initialize an Excel file with proper headers and formatting for health records"""
    if excel_file is None:
        excel_file = current_app.config['EXCEL_FILE']
    
    # Get the standard headers
    headers = get_excel_headers()
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(excel_file), exist_ok=True)
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    sheet = wb.create_sheet('Health Records')
    
    # Write headers
    sheet.append(headers)
    
    # Style: headers bold, colored, centered; freeze top row; set row height
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.freeze_panes = 'A2'
    sheet.row_dimensions[1].height = 22
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths based on header length as initial spacing
    for col in sheet.columns:
        try:
            column = col[0].column_letter
        except Exception:
            continue
        header_len = len(str(col[0].value)) if col and col[0].value else 10
        adjusted_width = max(12, min((header_len + 2) * 1.1, 32))
        sheet.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(excel_file)
    return True

def format_value_for_excel(value):
    """Format values for proper Excel display"""
    if value is None:
        return ''
    elif isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    elif isinstance(value, time):
        return value.strftime('%H:%M')
    elif isinstance(value, bool):
        return 'Yes' if value else 'No'
    return str(value)

def export_records_to_excel(records, output_file=None):
    """Export database records to Excel file with proper formatting"""
    if not records:
        return None
    
    if output_file is None:
        output_file = os.path.join(
            os.path.dirname(current_app.config['EXCEL_FILE']),
            f"nurse_records_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    ws = wb.create_sheet('Health Records')
    
    # Get headers and write them to the worksheet
    headers = get_excel_headers()
    ws.append(headers)
    
    # Apply header styling, freeze pane, and initial widths
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for record in records:
        # Normalize to dict
        if hasattr(record, '__table__'):
            record = {c.name: getattr(record, c.name) for c in record.__table__.columns}
        
        # Row assembly with header-to-field mapping
        row = []
        for header in headers:
            # Default mapping based on header
            normalized = header.lower().strip()
            field_name = normalized.replace(' ', '_')
            
            # Explicit mappings for discrepancies/special cases
            if header == 'Parent/Guardian Name':
                field_name = 'parent_primary_name'
            elif header == 'Parent/Guardian Phone':
                field_name = 'parent_primary_phone'
            elif header == 'Emergency Contact Name':
                field_name = 'emergency_contact_name'
            elif header == 'Emergency Contact Phone':
                field_name = 'emergency_contact_phone'
            elif header == 'Grade/Year Level':
                field_name = 'grade_level'
            elif header == 'Visit Reason Category':
                # Prefer new column, fallback to legacy
                value = record.get('visit_reason_category') or record.get('visit_reason')
                row.append(format_value_for_excel(value))
                continue
            elif header == 'Temperature (°C)':
                field_name = 'temperature'
            elif header == 'Pulse (bpm)':
                field_name = 'heart_rate'
            elif header == 'Respiratory Rate (cpm)':
                field_name = 'respiratory_rate'
            elif header == 'Oxygen Saturation (%)':
                field_name = 'oxygen_saturation'
            elif header == 'Blood Pressure (mmHg)':
                # Combine systolic/diastolic if available
                sys_v = record.get('blood_pressure_systolic')
                dia_v = record.get('blood_pressure_diastolic')
                if sys_v is not None and dia_v is not None:
                    row.append(f"{sys_v}/{dia_v}")
                else:
                    # Fallback to any legacy combined field
                    row.append(format_value_for_excel(record.get('blood_pressure')))
                continue
            elif header == 'Visit Details':
                field_name = 'visit_details'
            elif header == 'Special Medical Needs':
                field_name = 'special_medical_needs'
            elif header == 'Next Step(s)':
                field_name = 'next_steps'
            elif header == 'Parent Notified':
                field_name = 'parent_notified'
            elif header == 'Incident Report Required':
                field_name = 'incident_report_required'
            elif header == 'Return to Class Time':
                field_name = 'return_to_class_time'
            elif header == 'Brought In By':
                field_name = 'brought_in_by'
            
            # Get the value and format it
            value = record.get(field_name, '')
            row.append(format_value_for_excel(value))
        
        ws.append(row)
    
    # Auto-size columns based on content (bounded), and set alignment for cells
    max_width = 32
    min_width = 12
    center_headers = {
        'Date of Visit', 'Time of Visit', 'Temperature (°C)', 'Pulse (bpm)',
        'Respiratory Rate (cpm)', 'Oxygen Saturation (%)', 'Blood Pressure (mmHg)',
        'Grade/Year Level', 'Parent Notified', 'Incident Report Required'
    }
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for col in ws.columns:
        try:
            column_letter = col[0].column_letter
        except Exception:
            continue
        # Determine width
        lengths = []
        for cell in col:
            val = '' if cell.value is None else str(cell.value)
            lengths.append(len(val))
            # Apply alignment: center for numeric/vitals/dates headers, left otherwise
            header_text = ws.cell(row=1, column=cell.column).value
            if cell.row == 1:
                continue
            if header_text in center_headers:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        best = max(lengths or [min_width])
        ws.column_dimensions[column_letter].width = max(min_width, min(int(best * 1.1) + 2, max_width))
    
    # Save the workbook
    wb.save(output_file)
    return output_file
